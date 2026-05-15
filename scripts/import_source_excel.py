from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROVINCE_CODES = {
    "北京": "110000",
    "北京市": "110000",
    "天津": "120000",
    "天津市": "120000",
    "河北": "130000",
    "河北省": "130000",
    "山西": "140000",
    "山西省": "140000",
    "内蒙古": "150000",
    "内蒙古自治区": "150000",
    "辽宁": "210000",
    "辽宁省": "210000",
    "吉林": "220000",
    "吉林省": "220000",
    "黑龙江": "230000",
    "黑龙江省": "230000",
    "上海": "310000",
    "上海市": "310000",
    "江苏": "320000",
    "江苏省": "320000",
    "浙江": "330000",
    "浙江省": "330000",
    "安徽": "340000",
    "安徽省": "340000",
    "福建": "350000",
    "福建省": "350000",
    "江西": "360000",
    "江西省": "360000",
    "山东": "370000",
    "山东省": "370000",
    "河南": "410000",
    "河南省": "410000",
    "湖北": "420000",
    "湖北省": "420000",
    "湖南": "430000",
    "湖南省": "430000",
    "广东": "440000",
    "广东省": "440000",
    "广西": "450000",
    "广西壮族自治区": "450000",
    "海南": "460000",
    "海南省": "460000",
    "重庆": "500000",
    "重庆市": "500000",
    "四川": "510000",
    "四川省": "510000",
    "贵州": "520000",
    "贵州省": "520000",
    "云南": "530000",
    "云南省": "530000",
    "西藏": "540000",
    "西藏自治区": "540000",
    "陕西": "610000",
    "陕西省": "610000",
    "甘肃": "620000",
    "甘肃省": "620000",
    "青海": "630000",
    "青海省": "630000",
    "宁夏": "640000",
    "宁夏回族自治区": "640000",
    "新疆": "650000",
    "新疆维吾尔自治区": "650000",
}


SUBJECT_CATEGORY_CODES = {
    "物理": "physics",
    "历史": "history",
    "理科": "science",
    "文科": "arts",
}


EDUCATION_LEVEL_CODES = {
    "本科": "undergraduate",
    "专科": "specialist",
}


SCHOOL_OWNERSHIP_CODES = {
    "公办": "public",
    "民办": "private",
    "中外合作办学": "sino_foreign",
    "内地与港澳台地区合作办学": "mainland_hmt_joint",
}


SCHOOL_CATEGORY_CODES = {
    "综合类": "comprehensive",
    "理工类": "science_engineering",
    "医药类": "medicine",
    "师范类": "normal",
    "财经类": "finance_economics",
    "农林类": "agriculture_forestry",
    "语言类": "language",
    "政法类": "politics_law",
    "艺术类": "art",
    "体育类": "sports",
    "民族类": "ethnic",
    "军事类": "military",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def int_or_empty(value: Any) -> str:
    text = clean(value).replace(",", "")
    if not text or text == "/":
        return ""
    match = re.search(r"-?\d+", text)
    return match.group(0) if match else ""


def percent_or_empty(value: Any) -> str:
    text = clean(value)
    if not text or text == "/":
        return ""
    text = text.replace("%", "")
    try:
        return str(round(float(text), 2))
    except ValueError:
        return ""


def decimal_or_empty(value: Any) -> str:
    text = clean(value)
    if not text or text == "/":
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    return match.group(0) if match else ""


def yes_no(value: Any) -> str:
    text = clean(value)
    if text == "是":
        return "true"
    if text == "否":
        return "false"
    return ""


def code_from_name(mapping: dict[str, str], name: str, prefix: str) -> str:
    if not name:
        return ""
    if name in mapping:
        return mapping[name]
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{digest}"


def province_code(name: str) -> str:
    if not name:
        return ""
    return PROVINCE_CODES.get(name, PROVINCE_CODES.get(name.replace("省", "").replace("市", ""), ""))


def batch_code(name: str) -> str:
    if not name:
        return ""
    special = {
        "普通本科批": "regular_undergraduate",
        "本科提前批": "early_undergraduate",
        "专科批": "specialist",
    }
    return special.get(name, code_from_name({}, name, "batch"))


def subject_requirement_code(name: str) -> str:
    if not name:
        return "none"
    special = {
        "不限": "none",
        "化": "chemistry",
        "化学": "chemistry",
        "物+化": "physics_chemistry",
        "物理+化学": "physics_chemistry",
    }
    return special.get(name, code_from_name({}, name, "subject_req"))


def normalized_subjects(name: str) -> str:
    if name in ("", "不限"):
        return "[]"
    subjects = []
    if "物" in name:
        subjects.append("物理")
    if "化" in name:
        subjects.append("化学")
    if "生" in name:
        subjects.append("生物")
    if "政" in name:
        subjects.append("政治")
    if "史" in name:
        subjects.append("历史")
    if "地" in name:
        subjects.append("地理")
    return json.dumps(subjects, ensure_ascii=False)


@dataclass(frozen=True)
class RowContext:
    university_code: str
    university_name: str
    batch_name: str
    education_level_name: str
    subject_category_name: str
    group_code: str
    subject_requirement_name: str
    local_major_code: str
    local_major_name_2025: str


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def run(command: list[str], cwd: Path | None = None) -> None:
    subprocess.run(command, cwd=cwd, check=True)


def psql(sql: str) -> None:
    run(["docker", "exec", "-i", "admission-db", "psql", "-U", "app", "-d", "admission", "-v", "ON_ERROR_STOP=1"], None)


def copy_and_import(tmp: Path, sql_path: Path) -> None:
    container_dir = "/tmp/source_import"
    run(["docker", "exec", "admission-db", "rm", "-rf", container_dir])
    run(["docker", "exec", "admission-db", "mkdir", "-p", container_dir])
    for file in tmp.glob("*.csv"):
        run(["docker", "cp", str(file), f"admission-db:{container_dir}/{file.name}"])
    run(["docker", "cp", str(sql_path), f"admission-db:{container_dir}/import.sql"])
    run(["docker", "exec", "admission-db", "psql", "-U", "app", "-d", "admission", "-v", "ON_ERROR_STOP=1", "-f", f"{container_dir}/import.sql"])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--sql", required=True)
    parser.add_argument("--keep-temp", action="store_true")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    sql_path = Path(args.sql)
    tmp = Path(tempfile.mkdtemp(prefix="source_import_"))
    print(f"Using temp dir: {tmp}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    idx = {name: pos for pos, name in enumerate(headers)}

    def cell(row: tuple[Any, ...], name: str) -> str:
        return clean(row[idx[name]]) if name in idx and idx[name] < len(row) else ""

    dictionaries: dict[str, dict[str, str]] = {
        "regions": {"230000": "黑龙江省", "110000": "北京市"},
        "subject_categories": {v: k for k, v in SUBJECT_CATEGORY_CODES.items()},
        "subject_requirements": {
            "none": "不限",
            "chemistry": "化学",
            "physics_chemistry": "物理+化学",
        },
        "batches": {
            "regular_undergraduate": "普通本科批",
            "early_undergraduate": "本科提前批",
            "specialist": "专科批",
        },
        "education_levels": {v: k for k, v in EDUCATION_LEVEL_CODES.items()},
        "school_ownership_types": {v: k for k, v in SCHOOL_OWNERSHIP_CODES.items()},
        "school_categories": {v: k for k, v in SCHOOL_CATEGORY_CODES.items()},
    }
    subject_requirement_subjects: dict[str, str] = {
        "none": "[]",
        "chemistry": '["化学"]',
        "physics_chemistry": '["物理", "化学"]',
    }

    universities: dict[tuple[str, str], dict[str, Any]] = {}
    profiles: dict[tuple[str, str, int], dict[str, Any]] = {}
    groups: dict[tuple[str, str, int, str, str, str, str], dict[str, Any]] = {}
    group_extensions: dict[tuple[str, str, int, str, str, str, str], dict[str, Any]] = {}
    admissions: dict[tuple[str, str, int, str, str, str, str, str], dict[str, Any]] = {}
    major_profiles: dict[tuple[str, str, int, str, str, str, str, str], dict[str, Any]] = {}
    postgraduate_profiles: dict[tuple[str, str, int], dict[str, Any]] = {}

    admissions_region_code = "230000"
    profile_year = 2025

    # 招生人数列：2024 xlsx 用 "录取人数1"（实际录取数据），其他年份用 "计划人数X"
    # 或 "25年计划人数"。schema 统一成 admitted_count 一个字段。
    year_specs = [
        {
            "year": 2025,
            "name_field": "25年招生计划专业名称",
            "admitted_field": "25年计划人数",
            "min_score_field": "",
            "min_rank_field": "",
            "max_score_field": "",
            "max_rank_field": "",
            "equiv_field": "",
            "include_plan_meta": True,
        },
        {
            "year": 2024,
            "name_field": "24年分数线专业名称",
            "admitted_field": "录取人数1",
            "min_score_field": "最低分1",
            "min_rank_field": "最低位次1",
            "max_score_field": "最高分1",
            "max_rank_field": "最高位次1",
            "equiv_field": "等位最低分24",
            "include_plan_meta": False,
        },
        {
            "year": 2023,
            "name_field": "23年分数线专业名称",
            "admitted_field": "计划人数2",
            "min_score_field": "最低分2",
            "min_rank_field": "最低位次2",
            "max_score_field": "最高分2",
            "max_rank_field": "最高位次2",
            "equiv_field": "等位最低分23",
            "include_plan_meta": False,
        },
        {
            "year": 2022,
            "name_field": "22年分数线专业名称",
            "admitted_field": "计划人数3",
            "min_score_field": "最低分3",
            "min_rank_field": "最低位次3",
            "max_score_field": "最高分3",
            "max_rank_field": "最高位次3",
            "equiv_field": "等位最低分22",
            "include_plan_meta": False,
        },
    ]

    for raw_row in ws.iter_rows(min_row=3, values_only=True):
        if not any(raw_row):
            continue
        university_code = cell(raw_row, "院校代码")
        university_name = cell(raw_row, "院校名称")
        group_code = cell(raw_row, "专业组")
        local_major_code = cell(raw_row, "专业代码")
        if not university_code or not university_name or not group_code or not local_major_code:
            continue

        batch_name = cell(raw_row, "批次")
        education_level_name = cell(raw_row, "本科/专科")
        subject_category_name = cell(raw_row, "科类")
        subject_requirement_name = cell(raw_row, "选科要求") or "不限"
        batch = batch_code(batch_name)
        education_level = code_from_name(EDUCATION_LEVEL_CODES, education_level_name, "edu")
        subject_category = code_from_name(SUBJECT_CATEGORY_CODES, subject_category_name, "subject")
        subject_requirement = subject_requirement_code(subject_requirement_name)

        dictionaries["batches"][batch] = batch_name
        dictionaries["education_levels"][education_level] = education_level_name
        dictionaries["subject_categories"][subject_category] = subject_category_name
        dictionaries["subject_requirements"][subject_requirement] = subject_requirement_name
        subject_requirement_subjects[subject_requirement] = normalized_subjects(subject_requirement_name)

        ownership_name = cell(raw_row, "办学性质")
        ownership = code_from_name(SCHOOL_OWNERSHIP_CODES, ownership_name, "ownership")
        if ownership_name:
            dictionaries["school_ownership_types"][ownership] = ownership_name

        school_category_name = cell(raw_row, "学校类别")
        school_category = code_from_name(SCHOOL_CATEGORY_CODES, school_category_name, "school_cat")
        if school_category_name:
            dictionaries["school_categories"][school_category] = school_category_name

        school_region_name = cell(raw_row, "院校省份")
        school_region = province_code(school_region_name)
        if school_region and school_region_name:
            dictionaries["regions"][school_region] = school_region_name if school_region_name.endswith(("省", "市", "自治区")) else school_region_name

        universities[(university_code, university_name)] = {
            "university_code": university_code,
            "name": university_name,
            "normalized_name": university_name.replace(" ", "").replace("（", "(").replace("）", ")"),
        }

        profiles[(university_code, university_name, profile_year)] = {
            "university_code": university_code,
            "university_name": university_name,
            "profile_year": profile_year,
            "region_code": school_region,
            "city": cell(raw_row, "院校城市"),
            "ownership_type_code": ownership,
            "school_category_code": school_category,
            "education_level_code": education_level,
            "is_985": yes_no(cell(raw_row, "985")),
            "is_211": yes_no(cell(raw_row, "211")),
            "is_double_first_class": yes_no(cell(raw_row, "双一流")),
            "is_national_key": yes_no(cell(raw_row, "国重点")),
            "is_provincial_key": yes_no(cell(raw_row, "省重点")),
            "has_postgraduate_recommendation": yes_no(cell(raw_row, "保研")),
            "postgraduate_recommendation_rate": percent_or_empty(cell(raw_row, "保研率")),
            "soft_rank": cell(raw_row, "院校软科排名"),
            "alumni_rank": cell(raw_row, "校友会排名"),
            "difficulty_rank": cell(raw_row, "录取难易度排名"),
            "doctoral_program_count": int_or_empty(cell(raw_row, "博士点数量")),
            "master_program_count": int_or_empty(cell(raw_row, "硕士点数量")),
            "national_key_subject_count": int_or_empty(cell(raw_row, "国家重点学科")),
            "affiliation": cell(raw_row, "学校归属"),
            "school_level_tags": cell(raw_row, "院校层次"),
            "excellence_tags": cell(raw_row, "部属卓越拔尖"),
        }
        postgraduate_profiles[(university_code, university_name, profile_year)] = {
            "university_code": university_code,
            "university_name": university_name,
            "profile_year": profile_year,
            "master_major_count": int_or_empty(cell(raw_row, "硕士专业数量")),
            "master_major_names": cell(raw_row, "硕士专业"),
            "doctoral_major_count": int_or_empty(cell(raw_row, "博士专业数量")),
            "doctoral_major_names": cell(raw_row, "博士专业"),
        }

        for spec in year_specs:
            year = int(spec["year"])
            local_major_name = cell(raw_row, spec["name_field"])
            if not local_major_name and year == 2025:
                local_major_name = cell(raw_row, "25年招生计划专业名称")
            if not local_major_name:
                continue

            has_numbers = any(
                int_or_empty(cell(raw_row, field))
                for field in (
                    spec["admitted_field"],
                    spec["min_score_field"],
                    spec["min_rank_field"],
                    spec["max_score_field"],
                    spec["max_rank_field"],
                    spec["equiv_field"],
                )
                if field
            )
            if year != 2025 and not has_numbers:
                continue

            group_key = (
                university_code,
                university_name,
                year,
                admissions_region_code,
                subject_category,
                batch,
                group_code,
            )
            groups[group_key] = {
                "university_code": university_code,
                "university_name": university_name,
                "admission_year": year,
                "region_code": admissions_region_code,
                "subject_category_code": subject_category,
                "batch_code": batch,
                "group_code": group_code,
                "subject_requirement_code": subject_requirement,
                "education_level_code": education_level,
                "group_major_count": int_or_empty(cell(raw_row, "组内数量")),
                "group_major_names": cell(raw_row, "组内专业"),
                "group_type": cell(raw_row, "专业组类型"),
            }
            group_extensions[group_key] = {
                "university_code": university_code,
                "university_name": university_name,
                "admission_year": year,
                "region_code": admissions_region_code,
                "subject_category_code": subject_category,
                "batch_code": batch,
                "group_code": group_code,
                "batch_remark": cell(raw_row, "批次备注"),
                "group_min_score": int_or_empty(cell(raw_row, "专业组最低分")) if spec["include_plan_meta"] else "",
                "group_min_rank": int_or_empty(cell(raw_row, "专业组最低位次")) if spec["include_plan_meta"] else "",
                "equivalent_min_score_2024": int_or_empty(cell(raw_row, "等位最低分24")),
                "equivalent_min_score_2023": int_or_empty(cell(raw_row, "等位最低分23")),
                "equivalent_min_score_2022": int_or_empty(cell(raw_row, "等位最低分22")),
                "subject_change_2024": cell(raw_row, "24年选科情况") if spec["include_plan_meta"] else "",
            }

            admission_key = group_key + (local_major_code,)
            if admission_key in admissions:
                continue

            admissions[admission_key] = {
                "university_code": university_code,
                "university_name": university_name,
                "admission_year": year,
                "region_code": admissions_region_code,
                "subject_category_code": subject_category,
                "batch_code": batch,
                "group_code": group_code,
                "local_major_code": local_major_code,
                "local_major_name": local_major_name,
                "admitted_count": int_or_empty(cell(raw_row, spec["admitted_field"])),
                "min_score": int_or_empty(cell(raw_row, spec["min_score_field"])),
                "min_rank": int_or_empty(cell(raw_row, spec["min_rank_field"])),
                "max_score": int_or_empty(cell(raw_row, spec["max_score_field"])),
                "max_rank": int_or_empty(cell(raw_row, spec["max_rank_field"])),
                "equivalent_min_score": int_or_empty(cell(raw_row, spec["equiv_field"])),
                "tuition": int_or_empty(cell(raw_row, "学费")) if spec["include_plan_meta"] else "",
                "duration": cell(raw_row, "学制") if spec["include_plan_meta"] else "",
                "admission_remark": cell(raw_row, "25年招生计划专业备注") if spec["include_plan_meta"] else "",
                "major_intro": cell(raw_row, "专业简介") if spec["include_plan_meta"] else "",
                "training_goal": cell(raw_row, "培养目标") if spec["include_plan_meta"] else "",
                "subject_study_requirement": cell(raw_row, "学科要求") if spec["include_plan_meta"] else "",
                "main_courses": cell(raw_row, "本科主要课程") if spec["include_plan_meta"] else "",
                "postgraduate_direction": cell(raw_row, "考研方向") if spec["include_plan_meta"] else "",
                "employment_direction": cell(raw_row, "就业方向") if spec["include_plan_meta"] else "",
            }
            if spec["include_plan_meta"]:
                major_profiles[admission_key] = {
                    "university_code": university_code,
                    "university_name": university_name,
                    "admission_year": year,
                    "region_code": admissions_region_code,
                    "subject_category_code": subject_category,
                    "batch_code": batch,
                    "group_code": group_code,
                    "local_major_code": local_major_code,
                    "discipline_category": cell(raw_row, "学科门类"),
                    "first_level_discipline": cell(raw_row, "一级学科"),
                    "fourth_round_subject_eval": cell(raw_row, "第四轮学科评估"),
                    "double_first_class_subject": cell(raw_row, "双一流学科"),
                    "soft_major_grade": cell(raw_row, "软科等级"),
                    "major_evaluation_score": decimal_or_empty(cell(raw_row, "专业评价得分")),
                    "major_rank": cell(raw_row, "专业排名"),
                    "is_national_feature": yes_no(cell(raw_row, "国家特色")),
                    "corresponding_master_majors": cell(raw_row, "本校对应硕士专业"),
                    "corresponding_doctoral_majors": cell(raw_row, "本校对应博士专业"),
                }

    csv_specs = [
        ("regions.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["regions"].items()) if k and v]),
        ("subject_categories.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["subject_categories"].items()) if k and v]),
        ("subject_requirements.csv", ["code", "name", "normalized_subjects"], [{"code": k, "name": v, "normalized_subjects": subject_requirement_subjects.get(k, "[]")} for k, v in sorted(dictionaries["subject_requirements"].items()) if k and v]),
        ("batches.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["batches"].items()) if k and v]),
        ("education_levels.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["education_levels"].items()) if k and v]),
        ("school_ownership_types.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["school_ownership_types"].items()) if k and v]),
        ("school_categories.csv", ["code", "name"], [{"code": k, "name": v} for k, v in sorted(dictionaries["school_categories"].items()) if k and v]),
        ("universities.csv", ["university_code", "name", "normalized_name"], list(universities.values())),
        ("university_profiles.csv", [
            "university_code", "university_name", "profile_year", "region_code", "city",
            "ownership_type_code", "school_category_code", "education_level_code",
            "is_985", "is_211", "is_double_first_class", "is_national_key", "is_provincial_key",
            "has_postgraduate_recommendation", "postgraduate_recommendation_rate", "soft_rank",
            "alumni_rank", "difficulty_rank", "doctoral_program_count", "master_program_count",
            "national_key_subject_count", "affiliation", "school_level_tags", "excellence_tags",
        ], list(profiles.values())),
        ("admission_groups.csv", [
            "university_code", "university_name", "admission_year", "region_code",
            "subject_category_code", "batch_code", "group_code", "subject_requirement_code",
            "education_level_code", "group_major_count", "group_major_names", "group_type",
        ], list(groups.values())),
        ("admission_group_extensions.csv", [
            "university_code", "university_name", "admission_year", "region_code",
            "subject_category_code", "batch_code", "group_code", "batch_remark",
            "group_min_score", "group_min_rank", "equivalent_min_score_2024",
            "equivalent_min_score_2023", "equivalent_min_score_2022", "subject_change_2024",
        ], list(group_extensions.values())),
        ("university_major_admissions.csv", [
            "university_code", "university_name", "admission_year", "region_code",
            "subject_category_code", "batch_code", "group_code", "local_major_code",
            "local_major_name", "admitted_count", "min_score", "min_rank",
            "max_score", "max_rank", "equivalent_min_score", "tuition", "duration",
            "admission_remark", "major_intro", "training_goal", "subject_study_requirement",
            "main_courses", "postgraduate_direction", "employment_direction",
        ], list(admissions.values())),
        ("university_major_profiles.csv", [
            "university_code", "university_name", "admission_year", "region_code",
            "subject_category_code", "batch_code", "group_code", "local_major_code",
            "discipline_category", "first_level_discipline", "fourth_round_subject_eval",
            "double_first_class_subject", "soft_major_grade", "major_evaluation_score",
            "major_rank", "is_national_feature", "corresponding_master_majors",
            "corresponding_doctoral_majors",
        ], list(major_profiles.values())),
        ("university_postgraduate_profiles.csv", [
            "university_code", "university_name", "profile_year", "master_major_count",
            "master_major_names", "doctoral_major_count", "doctoral_major_names",
        ], list(postgraduate_profiles.values())),
    ]

    for filename, fields, rows in csv_specs:
        write_csv(tmp / filename, fields, rows)
        print(f"{filename}: {len(rows)}")

    copy_and_import(tmp, sql_path)
    if not args.keep_temp:
        shutil.rmtree(tmp)


if __name__ == "__main__":
    main()
